#!/usr/bin/env python3

import requests
from openpyxl import load_workbook
from datetime import date

# Login
login_response = requests.post(
    'http://localhost:8000/token',
    data={'username': 'admin', 'password': 'admin123'}
)
token = login_response.json().get('access_token')

# Get weekly report
headers = {'Authorization': f'Bearer {token}'}
start_date = date(2025, 12, 19)
end_date = date(2025, 12, 25)

response = requests.get(
    f'http://localhost:8000/attendance/export/weekly?department_id=1&start_date={start_date}&end_date={end_date}',
    headers=headers
)

# Save
with open('/tmp/weekly_test.xlsx', 'wb') as f:
    f.write(response.content)

# Check sheets
wb = load_workbook('/tmp/weekly_test.xlsx')
print(f"Sheets in order: {wb.sheetnames}")
print(f"Active sheet: {wb.active.title}")
