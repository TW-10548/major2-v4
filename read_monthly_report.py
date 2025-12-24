#!/usr/bin/env python3

from openpyxl import load_workbook

# Load the report
wb = load_workbook('/tmp/monthly_report_debug.xlsx')

print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n\n=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    
    # Print first 10 rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        print(f"Row {row_idx}:", end=" ")
        for cell in row:
            if cell.value:
                print(f"[{cell.column_letter}{cell.row}: {cell.value}]", end=" ")
        print()

# Now check specifically the Attendance Details sheet
if 'Attendance Details' in wb.sheetnames:
    ws = wb['Attendance Details']
    print(f"\n\n=== Attendance Details Sheet (ALL DATA) ===")
    print(f"Max Row: {ws.max_row}")
    print(f"Max Column: {ws.max_column}")
    
    # Print all rows
    for row_idx in range(1, min(ws.max_row + 1, 50)):
        values = []
        for col_idx in range(1, min(ws.max_column + 1, 15)):
            cell = ws.cell(row=row_idx, column=col_idx)
            values.append(str(cell.value))
        print(f"Row {row_idx}: {values}")
