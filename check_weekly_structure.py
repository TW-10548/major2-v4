#!/usr/bin/env python3

import requests
from openpyxl import load_workbook
from datetime import date
import time

# Wait for server
time.sleep(2)

# Login
try:
    login_response = requests.post(
        'http://localhost:8000/token',
        data={'username': 'admin', 'password': 'admin123'},
        timeout=5
    )
    token = login_response.json().get('access_token')
except Exception as e:
    print(f"Login error: {e}")
    exit(1)

# Get weekly report
headers = {'Authorization': f'Bearer {token}'}
start_date = date(2025, 12, 19)
end_date = date(2025, 12, 25)

try:
    response = requests.get(
        f'http://localhost:8000/attendance/export/weekly?department_id=1&start_date={start_date}&end_date={end_date}',
        headers=headers,
        timeout=10
    )
    print(f"Response status: {response.status_code}, size: {len(response.content)}")
except Exception as e:
    print(f"Request error: {e}")
    exit(1)

# Save
filename = '/tmp/weekly_check_structure.xlsx'
with open(filename, 'wb') as f:
    f.write(response.content)

# Check sheets
try:
    wb = load_workbook(filename)
    print(f"\nWorkbook properties:")
    print(f"  Total sheets: {len(wb.sheetnames)}")
    print(f"  Sheet names: {wb.sheetnames}")
    print(f"  Active sheet: {wb.active.title}")
    
    # Check Summary sheet
    print(f"\n=== SUMMARY SHEET ===")
    summary = wb['Summary']
    print(f"  Dimensions: {summary.dimensions}")
    print(f"  Max row: {summary.max_row}")
    for row_idx in range(1, min(8, summary.max_row + 1)):
        val_a = summary[f'A{row_idx}'].value
        val_b = summary[f'B{row_idx}'].value
        if val_a or val_b:
            print(f"    Row {row_idx}: {val_a} | {val_b}")
    
    # Check Attendance Details sheet
    print(f"\n=== ATTENDANCE DETAILS SHEET ===")
    details = wb['Attendance Details']
    print(f"  Dimensions: {details.dimensions}")
    print(f"  Max row: {details.max_row}")
    print(f"  Data row sample (row 5):")
    print(f"    {details['A5'].value} | {details['B5'].value} | {details['C5'].value}")
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
