#!/usr/bin/env python3

import requests
from openpyxl import load_workbook

# Login
login_response = requests.post(
    'http://localhost:8000/token',
    data={'username': 'admin', 'password': 'admin123'}
)
token = login_response.json().get('access_token')

# Get monthly report
headers = {'Authorization': f'Bearer {token}'}
response = requests.get(
    'http://localhost:8000/attendance/export/monthly?department_id=1&year=2025&month=12',
    headers=headers
)

# Save
with open('/tmp/monthly_test.xlsx', 'wb') as f:
    f.write(response.content)

# Check sheets
wb = load_workbook('/tmp/monthly_test.xlsx')
print(f"Sheets in order: {wb.sheetnames}")
print(f"Active sheet: {wb.active.title}")
