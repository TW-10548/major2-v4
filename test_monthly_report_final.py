#!/usr/bin/env python3

import requests

# Login as admin using OAuth2 form data
login_response = requests.post(
    'http://localhost:8000/token',
    data={'username': 'admin', 'password': 'admin123'}
)

if login_response.status_code != 200:
    print(f"Login failed: {login_response.status_code}")
    exit(1)

token = login_response.json().get('access_token')

# Test monthly export
headers = {'Authorization': f'Bearer {token}'}
params = {'department_id': 1, 'year': 2025, 'month': 12}

print(f"Fetching monthly report...")
response = requests.get(
    'http://localhost:8000/attendance/export/monthly',
    headers=headers,
    params=params
)

print(f"Status code: {response.status_code}")
print(f"Content size: {len(response.content)} bytes")

# Save the file
with open('/tmp/monthly_report_final.xlsx', 'wb') as f:
    f.write(response.content)

# Read and display the sheets
from openpyxl import load_workbook

wb = load_workbook('/tmp/monthly_report_final.xlsx')
print(f"\nSheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    
    if sheet_name == "Summary":
        # Show summary stats
        for row_idx in range(1, min(ws.max_row + 1, 15)):
            values = []
            for col_idx in range(1, 3):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    values.append(str(cell.value)[:40])
            if values:
                print(f"Row {row_idx}: {values}")
    else:
        # Show first few attendance records
        print(f"Total rows: {ws.max_row}")
        for row_idx in range(1, min(ws.max_row + 1, 8)):
            values = []
            for col_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=col_idx)
                values.append(str(cell.value)[:25] if cell.value else "")
            print(f"Row {row_idx}: {values}")
