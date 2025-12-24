#!/usr/bin/env python3

import requests
from openpyxl import load_workbook
import time

# Wait for server
time.sleep(2)

# Login
try:
    login_response = requests.post(
        'http://localhost:8000/token',
        data={'username': 'admin', 'password': 'admin123'},
        timeout=5
    )
    token = login_response.json().get('access_token')
except Exception as e:
    print(f"Login error: {e}")
    exit(1)

# Get monthly report
headers = {'Authorization': f'Bearer {token}'}
try:
    response = requests.get(
        'http://localhost:8000/attendance/export/monthly?department_id=1&year=2025&month=12',
        headers=headers,
        timeout=10
    )
    print(f"Response status: {response.status_code}, size: {len(response.content)}")
except Exception as e:
    print(f"Request error: {e}")
    exit(1)

# Save
with open('/tmp/test_both_sheets.xlsx', 'wb') as f:
    f.write(response.content)

# Check sheets
try:
    wb = load_workbook('/tmp/test_both_sheets.xlsx')
    print(f"\nTotal sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")
    print(f"Active sheet: {wb.active.title}")
    
    # Check Summary sheet
    if 'Summary' in wb.sheetnames:
        summary = wb['Summary']
        print(f"\nSummary sheet has {summary.max_row} rows")
        print("First 10 rows of Summary:")
        for row_idx in range(1, min(11, summary.max_row + 1)):
            val_a = summary[f'A{row_idx}'].value
            val_b = summary[f'B{row_idx}'].value
            if val_a or val_b:
                print(f"  Row {row_idx}: {val_a} | {val_b}")
    else:
        print("\nSummary sheet NOT FOUND")
    
    # Check Details sheet
    if 'Attendance Details' in wb.sheetnames:
        details = wb['Attendance Details']
        print(f"\nAttendance Details sheet has {details.max_row} rows")
        print(f"Active cell: {details.active_cell}")
    else:
        print("\nAttendance Details sheet NOT FOUND")
        
except Exception as e:
    print(f"Excel read error: {e}")
    import traceback
    traceback.print_exc()
