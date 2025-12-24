#!/usr/bin/env python3

import requests
from datetime import date, timedelta

# Login as admin using OAuth2 form data
login_response = requests.post(
    'http://localhost:8000/token',
    data={'username': 'admin', 'password': 'admin123'}
)

if login_response.status_code != 200:
    print(f"Login failed: {login_response.status_code}")
    print(login_response.text)
    exit(1)

token = login_response.json().get('access_token')
print(f"Token obtained: {token[:20]}...")

# Test weekly export with dates
headers = {'Authorization': f'Bearer {token}'}

# Get the week of Dec 19-25, 2025
start_date = date(2025, 12, 19)
end_date = date(2025, 12, 25)

params = {
    'department_id': 1,
    'start_date': start_date.isoformat(),
    'end_date': end_date.isoformat()
}

print(f"\nFetching weekly report with params: {params}")
response = requests.get(
    'http://localhost:8000/attendance/export/weekly',
    headers=headers,
    params=params
)

print(f"Status code: {response.status_code}")
print(f"Content type: {response.headers.get('content-type')}")
print(f"Content size: {len(response.content)} bytes")

# Save the file
with open('/tmp/weekly_report_test.xlsx', 'wb') as f:
    f.write(response.content)

print(f"\nFile saved to /tmp/weekly_report_test.xlsx")

# Now read and display the sheets
from openpyxl import load_workbook

wb = load_workbook('/tmp/weekly_report_test.xlsx')
print(f"\nSheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    
    # Print first 15 rows
    for row_idx in range(1, min(ws.max_row + 1, 15)):
        values = []
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            values.append(str(cell.value)[:30] if cell.value else "")
        print(f"Row {row_idx}: {values}")
