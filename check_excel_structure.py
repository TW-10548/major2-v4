#!/usr/bin/env python3

import requests
from openpyxl import load_workbook
import time

# Wait for server
time.sleep(2)

# Login
try:
    login_response = requests.post(
        'http://localhost:8000/token',
        data={'username': 'admin', 'password': 'admin123'},
        timeout=5
    )
    token = login_response.json().get('access_token')
except Exception as e:
    print(f"Login error: {e}")
    exit(1)

# Get monthly report
headers = {'Authorization': f'Bearer {token}'}
try:
    response = requests.get(
        'http://localhost:8000/attendance/export/monthly?department_id=1&year=2025&month=12',
        headers=headers,
        timeout=10
    )
    print(f"Response status: {response.status_code}, size: {len(response.content)}")
except Exception as e:
    print(f"Request error: {e}")
    exit(1)

# Save
filename = '/tmp/monthly_check_structure.xlsx'
with open(filename, 'wb') as f:
    f.write(response.content)

# Check sheets with detailed info
try:
    wb = load_workbook(filename)
    print(f"\nWorkbook properties:")
    print(f"  Total sheets: {len(wb.sheetnames)}")
    print(f"  Sheet names: {wb.sheetnames}")
    print(f"  Active sheet index: {wb.index(wb.active)}")
    print(f"  Active sheet name: {wb.active.title}")
    
    # Check Summary sheet
    print(f"\n=== SUMMARY SHEET ===")
    summary = wb['Summary']
    print(f"  Dimensions: {summary.dimensions}")
    print(f"  Max row: {summary.max_row}")
    
    # Check Attendance Details sheet
    print(f"\n=== ATTENDANCE DETAILS SHEET ===")
    if 'Attendance Details' in wb.sheetnames:
        details = wb['Attendance Details']
        print(f"  Dimensions: {details.dimensions}")
        print(f"  Max row: {details.max_row}")
        print(f"  First 5 data rows:")
        for row_idx in range(5, min(10, details.max_row + 1)):
            emp_id = details[f'A{row_idx}'].value
            emp_name = details[f'B{row_idx}'].value
            emp_date = details[f'C{row_idx}'].value
            check_in = details[f'G{row_idx}'].value
            print(f"    Row {row_idx}: {emp_id} | {emp_name} | {emp_date} | Check-In: {check_in}")
    else:
        print("  ATTENDANCE DETAILS SHEET NOT FOUND!")
        
    print(f"\nFile saved to: {filename}")
    print(f"File size: {len(response.content)} bytes")
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
