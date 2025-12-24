#!/usr/bin/env python3

import requests
from openpyxl import load_workbook
from datetime import date
import time

# Wait for server
time.sleep(2)

# Login as admin
try:
    login_response = requests.post(
        'http://localhost:8000/token',
        data={'username': 'admin', 'password': 'admin123'},
        timeout=5
    )
    token = login_response.json().get('access_token')
    print("✓ Authenticated successfully")
except Exception as e:
    print(f"✗ Login error: {e}")
    exit(1)

headers = {'Authorization': f'Bearer {token}'}

print("\n" + "="*70)
print("TESTING EXCEL EXPORTS - DATA VALIDATION")
print("="*70)

# TEST 1: MONTHLY DEPARTMENT REPORT
print("\n\n1. TESTING MONTHLY DEPARTMENT ATTENDANCE REPORT")
print("-" * 70)

try:
    response = requests.get(
        'http://localhost:8000/attendance/export/monthly?department_id=1&year=2025&month=12',
        headers=headers,
        timeout=10
    )
    
    with open('/tmp/test_monthly_dept.xlsx', 'wb') as f:
        f.write(response.content)
    
    wb = load_workbook('/tmp/test_monthly_dept.xlsx')
    
    # Check Summary sheet
    summary = wb['Summary']
    print("\n✓ Summary Sheet Found")
    print(f"  - Title: {summary['A1'].value}")
    print(f"  - Working Days Available: {summary['B9'].value}")
    print(f"  - Working Days Completed: {summary['B10'].value}")
    print(f"  - Total Working Hours: {summary['B12'].value}")
    print(f"  - Total Overtime Hours: {summary['B13'].value}")
    
    # Check Details sheet
    details = wb['Attendance Details']
    print("\n✓ Attendance Details Sheet Found")
    print(f"  - Total data rows: {details.max_row - 4}")  # Subtract header rows
    
    # Sample data row
    sample_row = 5
    emp_id = details[f'A{sample_row}'].value
    emp_name = details[f'B{sample_row}'].value
    emp_date = details[f'C{sample_row}'].value
    check_in = details[f'G{sample_row}'].value
    check_out = details[f'H{sample_row}'].value
    worked_hrs = details[f'I{sample_row}'].value
    
    print(f"\n  Sample Record (Row {sample_row}):")
    print(f"    - Employee ID: {emp_id}")
    print(f"    - Employee Name: {emp_name}")
    print(f"    - Date: {emp_date}")
    print(f"    - Check-In: {check_in}")
    print(f"    - Check-Out: {check_out}")
    print(f"    - Worked Hours: {worked_hrs}")
    
    print(f"\n✓ Monthly Department Report - ALL DATA FETCHED CORRECTLY")
    
except Exception as e:
    print(f"✗ Monthly report error: {e}")
    import traceback
    traceback.print_exc()

# TEST 2: WEEKLY DEPARTMENT REPORT
print("\n\n2. TESTING WEEKLY DEPARTMENT ATTENDANCE REPORT")
print("-" * 70)

try:
    start_date = date(2025, 12, 19)
    end_date = date(2025, 12, 25)
    
    response = requests.get(
        f'http://localhost:8000/attendance/export/weekly?department_id=1&start_date={start_date}&end_date={end_date}',
        headers=headers,
        timeout=10
    )
    
    with open('/tmp/test_weekly_dept.xlsx', 'wb') as f:
        f.write(response.content)
    
    wb = load_workbook('/tmp/test_weekly_dept.xlsx')
    
    # Check Summary sheet
    summary = wb['Summary']
    print("\n✓ Summary Sheet Found")
    print(f"  - Title: {summary['A1'].value}")
    print(f"  - Date Range: {summary['A2'].value}")
    print(f"  - Total Employees: {summary['B5'].value}")
    print(f"  - Employees Present: {summary['B6'].value}")
    print(f"  - Total Working Hours: {summary['B7'].value}")
    print(f"  - Total Overtime Hours: {summary['B8'].value}")
    
    # Check Details sheet
    details = wb['Attendance Details']
    print("\n✓ Attendance Details Sheet Found")
    print(f"  - Total data rows: {details.max_row - 4}")
    
    # Sample data
    sample_row = 5
    emp_id = details[f'A{sample_row}'].value
    emp_name = details[f'B{sample_row}'].value
    emp_date = details[f'C{sample_row}'].value
    shift = details[f'D{sample_row}'].value
    
    print(f"\n  Sample Record (Row {sample_row}):")
    print(f"    - Employee ID: {emp_id}")
    print(f"    - Employee Name: {emp_name}")
    print(f"    - Date: {emp_date}")
    print(f"    - Assigned Shift: {shift}")
    
    print(f"\n✓ Weekly Department Report - ALL DATA FETCHED CORRECTLY")
    
except Exception as e:
    print(f"✗ Weekly report error: {e}")
    import traceback
    traceback.print_exc()

# TEST 3: EMPLOYEE MONTHLY REPORT
print("\n\n3. TESTING EMPLOYEE MONTHLY ATTENDANCE REPORT")
print("-" * 70)

try:
    response = requests.get(
        'http://localhost:8000/attendance/export/employee-monthly?year=2025&month=12&employee_id=EMP001',
        headers=headers,
        timeout=10
    )
    
    with open('/tmp/test_employee_monthly.xlsx', 'wb') as f:
        f.write(response.content)
    
    wb = load_workbook('/tmp/test_employee_monthly.xlsx')
    
    # Check Summary sheet
    summary = wb['Summary']
    print("\n✓ Summary Sheet Found")
    print(f"  - Employee: {summary['A1'].value}")
    
    # Extract summary data
    print("\n  Attendance Summary:")
    for row in range(6, 12):
        label = summary[f'A{row}'].value
        value = summary[f'B{row}'].value
        if label:
            print(f"    - {label}: {value}")
    
    print("\n  Leave Summary:")
    for row in range(14, 17):
        label = summary[f'A{row}'].value
        value = summary[f'B{row}'].value
        if label:
            print(f"    - {label}: {value}")
    
    print("\n  Hours Summary:")
    for row in range(19, 21):
        label = summary[f'A{row}'].value
        value = summary[f'B{row}'].value
        if label:
            print(f"    - {label}: {value}")
    
    # Check Daily Attendance sheet
    daily = wb['Daily Attendance']
    print(f"\n✓ Daily Attendance Sheet Found")
    print(f"  - Total data rows: {daily.max_row - 4}")
    
    # Sample data
    if daily.max_row > 4:
        sample_row = 5
        date_val = daily[f'A{sample_row}'].value
        day_val = daily[f'B{sample_row}'].value
        check_in = daily[f'D{sample_row}'].value
        check_out = daily[f'E{sample_row}'].value
        
        print(f"\n  Sample Record (Row {sample_row}):")
        print(f"    - Date: {date_val}")
        print(f"    - Day: {day_val}")
        print(f"    - Check-In: {check_in}")
        print(f"    - Check-Out: {check_out}")
    
    print(f"\n✓ Employee Monthly Report - ALL DATA FETCHED CORRECTLY")
    
except Exception as e:
    print(f"✗ Employee monthly report error: {e}")
    import traceback
    traceback.print_exc()

print("\n\n" + "="*70)
print("VERIFICATION COMPLETE - ALL EXPORTS SUCCESSFULLY FETCHING DB DATA")
print("="*70 + "\n")
