#!/usr/bin/env python3

import requests
from openpyxl import load_workbook
import time

# Wait for server
time.sleep(2)

# Login as admin
try:
    login_response = requests.post(
        'http://localhost:8000/token',
        data={'username': 'admin', 'password': 'admin123'},
        timeout=5
    )
    token = login_response.json().get('access_token')
except Exception as e:
    print(f"Login error: {e}")
    exit(1)

# Get employee monthly report for EMP001
headers = {'Authorization': f'Bearer {token}'}
try:
    response = requests.get(
        'http://localhost:8000/attendance/export/employee-monthly?year=2025&month=12&employee_id=EMP001',
        headers=headers,
        timeout=10
    )
    print(f"Response status: {response.status_code}, size: {len(response.content)}")
except Exception as e:
    print(f"Request error: {e}")
    exit(1)

# Save
filename = '/tmp/employee_monthly_report_test.xlsx'
with open(filename, 'wb') as f:
    f.write(response.content)

# Check sheets
try:
    wb = load_workbook(filename)
    print(f"\nWorkbook properties:")
    print(f"  Total sheets: {len(wb.sheetnames)}")
    print(f"  Sheet names: {wb.sheetnames}")
    print(f"  Active sheet: {wb.active.title}")
    
    # Check Summary sheet
    if 'Summary' in wb.sheetnames:
        summary = wb['Summary']
        print(f"\n=== SUMMARY SHEET ===")
        print(f"  Dimensions: {summary.dimensions}")
        for row_idx in range(1, min(25, summary.max_row + 1)):
            val_a = summary[f'A{row_idx}'].value
            val_b = summary[f'B{row_idx}'].value
            if val_a or val_b:
                print(f"  Row {row_idx}: {str(val_a)[:40]} | {str(val_b)[:20]}")
    
    # Check Daily Attendance sheet
    if 'Daily Attendance' in wb.sheetnames:
        daily = wb['Daily Attendance']
        print(f"\n=== DAILY ATTENDANCE SHEET ===")
        print(f"  Dimensions: {daily.dimensions}")
        print(f"  Sample data row 5:")
        for col in range(1, 5):
            print(f"    Col {col}: {daily.cell(5, col).value}")
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
